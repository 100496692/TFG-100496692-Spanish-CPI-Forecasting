"""
================================================================================
BACHELOR THESIS - SPANISH CPI FORECASTING
Data Collection, Cleaning, and Preprocessing Pipeline
Author: Álvaro
Date: May 2025
================================================================================
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Configuration
plt.style.use('default')
sns.set_palette("husl")
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 120)
pd.set_option('display.float_format', lambda x: '%.4f' % x)

print("="*80)
print("SPANISH CPI FORECASTING - DATA PREPROCESSING")
print("="*80)

#==============================================================================
# 1. LOAD AND CLEAN IPC (TARGET VARIABLE)
#==============================================================================

print("\n[1/7] Processing CPI (Índice de Precios al Consumo)...")

df_ipc_raw = pd.read_excel('FUENTE_GENERAL_IPC.xlsx', 
                           sheet_name='Series', 
                           header=None)

print(f"   Excel dimensions: {df_ipc_raw.shape[0]} rows × {df_ipc_raw.shape[1]} columns")

# Extract data: skip column 0 (empty), take columns 1-13 (Year + M01-M12)
df_ipc = df_ipc_raw.iloc[7:, 1:14].copy()
df_ipc.columns = ['Year'] + [f'M{i:02d}' for i in range(1, 13)]

# Convert Year to integer
df_ipc['Year'] = pd.to_numeric(df_ipc['Year'], errors='coerce')
df_ipc = df_ipc.dropna(subset=['Year'])
df_ipc['Year'] = df_ipc['Year'].astype(int)

print(f"   Years range: {df_ipc['Year'].min()} - {df_ipc['Year'].max()}")

# Melt to long format
df_ipc_long = df_ipc.melt(id_vars=['Year'], 
                          var_name='Month', 
                          value_name='CPI_Variation')

# Create date
df_ipc_long['Month_Num'] = df_ipc_long['Month'].str.extract('(\d+)').astype(int)
df_ipc_long['Date'] = pd.to_datetime(
    df_ipc_long['Year'].astype(str) + '-' + 
    df_ipc_long['Month_Num'].astype(str) + '-01'
)

# Filter 2003-2024
df_ipc_final = df_ipc_long[
    (df_ipc_long['Date'] >= '2003-01-01') & 
    (df_ipc_long['Date'] <= '2024-12-31')
][['Date', 'CPI_Variation']].copy()

df_ipc_final['CPI_Variation'] = pd.to_numeric(df_ipc_final['CPI_Variation'], errors='coerce')
df_ipc_final = df_ipc_final.sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_ipc_final)} obs | Range: {df_ipc_final['CPI_Variation'].min():.2f} to {df_ipc_final['CPI_Variation'].max():.2f}")

#==============================================================================
# 2. EURIBOR - DETECCIÓN AUTOMÁTICA DE FORMATO
#==============================================================================

print("\n[2/7] Processing Euribor 12M...")

df_euribor_raw = pd.read_excel('EURIBOR_MENSUAL.xlsx', 
                               sheet_name='Hoja1',
                               header=None,
                               skiprows=3)

df_euribor = df_euribor_raw[[0, 1]].copy()
df_euribor.columns = ['Date', 'Euribor_Raw']
df_euribor['Date'] = pd.to_datetime(df_euribor['Date'], errors='coerce')
df_euribor['Euribor_Raw'] = pd.to_numeric(df_euribor['Euribor_Raw'], errors='coerce')

# DETECCIÓN AUTOMÁTICA DE FORMATO
# Si valor absoluto > 10: dividir entre 1000 (formato antiguo: 2705 → 2.705)
# Si valor absoluto <= 10: ya está en porcentaje (0.877 → 0.877)
def convert_euribor(value):
    if pd.isna(value):
        return np.nan
    elif abs(value) > 10:
        # Formato antiguo: milésimas (2705 → 2.705%)
        return value / 1000
    else:
        # Formato nuevo: ya en porcentaje
        return value

df_euribor['Euribor_12M'] = df_euribor['Euribor_Raw'].apply(convert_euribor)

# Eliminar filas con valores no numéricos
df_euribor = df_euribor[df_euribor['Euribor_12M'].notna()].copy()

df_euribor_final = df_euribor[
    (df_euribor['Date'] >= '2003-01-01') & 
    (df_euribor['Date'] <= '2024-12-31')
][['Date', 'Euribor_12M']].copy().sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_euribor_final)} obs | Range: {df_euribor_final['Euribor_12M'].min():.3f}% to {df_euribor_final['Euribor_12M'].max():.3f}%")

#==============================================================================
# 3. FOOD CONSUMPTION - SIN CONVERSIÓN
#==============================================================================

print("\n[3/7] Processing Food Consumption...")

df_consumo_raw = pd.read_excel('CONSUMO_MENSUAL.xlsx', 
                               sheet_name='Hoja1',
                               header=None)

df_consumo = df_consumo_raw[[0, 1]].copy()
df_consumo.columns = ['Date', 'Food_Consumption']
df_consumo['Date'] = pd.to_datetime(df_consumo['Date'], errors='coerce')
df_consumo['Food_Consumption'] = pd.to_numeric(df_consumo['Food_Consumption'], errors='coerce')

df_consumo_final = df_consumo[
    (df_consumo['Date'] >= '2003-01-01') & 
    (df_consumo['Date'] <= '2024-12-31')
][['Date', 'Food_Consumption']].copy().sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_consumo_final)} obs | Range: {df_consumo_final['Food_Consumption'].min():.0f} to {df_consumo_final['Food_Consumption'].max():.0f}")

#==============================================================================
# 4. TOURISM
#==============================================================================

print("\n[4/7] Processing Tourism Overnight Stays...")

df_turismo_raw = pd.read_excel('TURISTASPERNOCTACIONES_MENSUAL.xlsx', 
                               sheet_name='tabla-2074',
                               header=None)

dates_row = df_turismo_raw.iloc[8, 1:].values
values_row = df_turismo_raw.iloc[9, 1:].values

df_turismo = pd.DataFrame({
    'Date_Str': dates_row,
    'Tourism_Overnight': values_row
})

df_turismo['Year'] = df_turismo['Date_Str'].str[:4]
df_turismo['Month'] = df_turismo['Date_Str'].str[5:7]
df_turismo['Date'] = pd.to_datetime(
    df_turismo['Year'] + '-' + df_turismo['Month'] + '-01',
    errors='coerce'
)
df_turismo['Tourism_Overnight'] = pd.to_numeric(df_turismo['Tourism_Overnight'], errors='coerce')

df_turismo_final = df_turismo[
    (df_turismo['Date'] >= '2003-01-01') & 
    (df_turismo['Date'] <= '2024-12-31')
][['Date', 'Tourism_Overnight']].copy().sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_turismo_final)} obs | Range: {df_turismo_final['Tourism_Overnight'].min():.0f} to {df_turismo_final['Tourism_Overnight'].max():.0f}")

#==============================================================================
# 5. MORTGAGES
#==============================================================================

print("\n[5/7] Processing Residential Mortgages...")

df_hipotecas_raw = pd.read_excel('HIPOTECAS_MENSUAL.xlsx', 
                                 sheet_name='tabla-13896',
                                 header=None)

dates_row = df_hipotecas_raw.iloc[8, 1:].values
values_row = df_hipotecas_raw.iloc[9, 1:].values

df_hipotecas = pd.DataFrame({
    'Date_Str': dates_row,
    'Mortgages': values_row
})

df_hipotecas['Year'] = df_hipotecas['Date_Str'].str[:4]
df_hipotecas['Month'] = df_hipotecas['Date_Str'].str[5:7]
df_hipotecas['Date'] = pd.to_datetime(
    df_hipotecas['Year'] + '-' + df_hipotecas['Month'] + '-01',
    errors='coerce'
)
df_hipotecas['Mortgages'] = pd.to_numeric(df_hipotecas['Mortgages'], errors='coerce')

df_hipotecas_final = df_hipotecas[
    (df_hipotecas['Date'] >= '2003-01-01') & 
    (df_hipotecas['Date'] <= '2024-12-31')
][['Date', 'Mortgages']].copy().sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_hipotecas_final)} obs | Range: {df_hipotecas_final['Mortgages'].min():.0f} to {df_hipotecas_final['Mortgages'].max():.0f}")

#==============================================================================
# 6. IPI
#==============================================================================

print("\n[6/7] Processing Industrial Production Index...")

df_ipi_raw = pd.read_excel('IPI_MENSUAL.xlsx', 
                           sheet_name='tabla-60272',
                           header=None)

dates_row = df_ipi_raw.iloc[7, 1:].values
values_row = df_ipi_raw.iloc[9, 1:].values

df_ipi = pd.DataFrame({
    'Date_Str': dates_row,
    'IPI': values_row
})

df_ipi['Year'] = df_ipi['Date_Str'].str[:4]
df_ipi['Month'] = df_ipi['Date_Str'].str[5:7]
df_ipi['Date'] = pd.to_datetime(
    df_ipi['Year'] + '-' + df_ipi['Month'] + '-01',
    errors='coerce'
)
df_ipi['IPI'] = pd.to_numeric(df_ipi['IPI'], errors='coerce')

df_ipi_final = df_ipi[
    (df_ipi['Date'] >= '2003-01-01') & 
    (df_ipi['Date'] <= '2024-12-31')
][['Date', 'IPI']].copy().sort_values('Date').reset_index(drop=True)

print(f"   ✓ Loaded: {len(df_ipi_final)} obs | Range: {df_ipi_final['IPI'].min():.2f} to {df_ipi_final['IPI'].max():.2f}")

#==============================================================================
# 7. EUR/USD - DETECCIÓN AUTOMÁTICA DE FORMATO + CORRECCIÓN
#==============================================================================

print("\n[7/7] Processing EUR/USD Exchange Rate...")

df_eurusd_raw = pd.read_excel('DOLAR_EURO_MENSUAL.xlsx', 
                              sheet_name='Hoja1',
                              header=None,
                              skiprows=3)

df_eurusd = df_eurusd_raw[[0, 1]].copy()
df_eurusd.columns = ['Date', 'EUR_USD_Raw']
df_eurusd['Date'] = pd.to_datetime(df_eurusd['Date'], errors='coerce')
df_eurusd['EUR_USD_Raw'] = pd.to_numeric(df_eurusd['EUR_USD_Raw'], errors='coerce')

# DETECCIÓN AUTOMÁTICA DE FORMATO
# Si > 100: dividir entre 10000 (formato antiguo: 10622 → 1.0622)
# Si < 100: ya está en decimal (formato nuevo: 0.99 → 0.99)
def convert_eurusd(value):
    if pd.isna(value):
        return np.nan
    elif value > 100:
        return value / 10000
    else:
        return value

df_eurusd['EUR_USD'] = df_eurusd['EUR_USD_Raw'].apply(convert_eurusd)

# CORRECCIÓN: Reemplazar valores anómalos (< 0.5 o > 2.0) con NaN
anomalies = ((df_eurusd['EUR_USD'] < 0.5) | (df_eurusd['EUR_USD'] > 2.0)) & (df_eurusd['EUR_USD'].notna())
anomalies_count = anomalies.sum()

if anomalies_count > 0:
    print(f"   ⚠️  {anomalies_count} anomalous values detected (outside 0.5-2.0 range)")
    df_eurusd.loc[anomalies, 'EUR_USD'] = np.nan

df_eurusd_final = df_eurusd[
    (df_eurusd['Date'] >= '2003-01-01') & 
    (df_eurusd['Date'] <= '2024-12-31')
][['Date', 'EUR_USD']].copy().sort_values('Date').reset_index(drop=True)

# Interpolar valores faltantes
df_eurusd_final['EUR_USD'] = df_eurusd_final['EUR_USD'].interpolate(method='linear')

print(f"   ✓ Loaded: {len(df_eurusd_final)} obs | Range: {df_eurusd_final['EUR_USD'].min():.4f} to {df_eurusd_final['EUR_USD'].max():.4f}")

#==============================================================================
# 8. MERGE
#==============================================================================

print("\n" + "="*80)
print("MERGING ALL DATASETS")
print("="*80)

df_master = df_ipc_final.copy()
df_master = df_master.merge(df_euribor_final, on='Date', how='left')
df_master = df_master.merge(df_consumo_final, on='Date', how='left')
df_master = df_master.merge(df_turismo_final, on='Date', how='left')
df_master = df_master.merge(df_hipotecas_final, on='Date', how='left')
df_master = df_master.merge(df_ipi_final, on='Date', how='left')
df_master = df_master.merge(df_eurusd_final, on='Date', how='left')

df_master.set_index('Date', inplace=True)

print(f"\n✓ Master dataset: {df_master.shape[0]} rows × {df_master.shape[1]} columns")
print(f"  Period: {df_master.index.min().strftime('%Y-%m')} to {df_master.index.max().strftime('%Y-%m')}")

#==============================================================================
# 9. MISSING VALUES
#==============================================================================

print("\n" + "="*80)
print("MISSING VALUES")
print("="*80)
print(df_master.isnull().sum())

df_clean = df_master.interpolate(method='linear', limit_direction='both')
df_clean = df_clean.fillna(method='ffill').fillna(method='bfill')

print("\n✓ All missing values handled")

#==============================================================================
# 10. FORMATO Y GUARDADO
#==============================================================================

print("\n" + "="*80)
print("APPLYING FORMAT AND SAVING")
print("="*80)

# Redondear con precisión adecuada
df_clean_formatted = df_clean.copy()
df_clean_formatted['CPI_Variation'] = df_clean_formatted['CPI_Variation'].round(2)
df_clean_formatted['Euribor_12M'] = df_clean_formatted['Euribor_12M'].round(3)  # 3 decimales
df_clean_formatted['Food_Consumption'] = df_clean_formatted['Food_Consumption'].round(2)
df_clean_formatted['Tourism_Overnight'] = df_clean_formatted['Tourism_Overnight'].round(0)
df_clean_formatted['Mortgages'] = df_clean_formatted['Mortgages'].round(0)
df_clean_formatted['IPI'] = df_clean_formatted['IPI'].round(3)
df_clean_formatted['EUR_USD'] = df_clean_formatted['EUR_USD'].round(4)

# Guardar pickle
df_clean_formatted.to_pickle('spanish_cpi_dataset_clean.pkl')
print("✓ Saved: spanish_cpi_dataset_clean.pkl")

# Guardar Excel con formato español
from openpyxl import load_workbook
from openpyxl.styles import numbers

df_clean_formatted.to_excel('spanish_cpi_dataset_clean.xlsx', engine='openpyxl')

wb = load_workbook('spanish_cpi_dataset_clean.xlsx')
ws = wb.active

# Formatos españoles
formato_decimal_2 = '#,##0.00'
formato_decimal_3 = '#,##0.000'
formato_decimal_4 = '#,##0.0000'
formato_entero = '#,##0'

# Aplicar formato por columna
for row in range(2, ws.max_row + 1):
    ws[f'B{row}'].number_format = formato_decimal_2  # CPI_Variation
    ws[f'C{row}'].number_format = formato_decimal_3  # Euribor_12M
    ws[f'D{row}'].number_format = formato_decimal_2  # Food_Consumption
    ws[f'E{row}'].number_format = formato_entero     # Tourism_Overnight
    ws[f'F{row}'].number_format = formato_entero     # Mortgages
    ws[f'G{row}'].number_format = formato_decimal_3  # IPI
    ws[f'H{row}'].number_format = formato_decimal_4  # EUR_USD

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12

wb.save('spanish_cpi_dataset_clean.xlsx')
print("✓ Saved: spanish_cpi_dataset_clean.xlsx (CON FORMATO ESPAÑOL)")

#==============================================================================
# 11. VALIDACIÓN FINAL
#==============================================================================

print("\n" + "="*80)
print("FINAL VALIDATION")
print("="*80)

for col in df_clean_formatted.columns:
    print(f"\n{col}:")
    print(f"  Min: {df_clean_formatted[col].min():.4f}")
    print(f"  Max: {df_clean_formatted[col].max():.4f}")
    print(f"  Mean: {df_clean_formatted[col].mean():.4f}")
    print(f"  Missing: {df_clean_formatted[col].isna().sum()}")

# Verificar EUR/USD específicamente
eur_low = (df_clean_formatted['EUR_USD'] < 0.8).sum()
eur_high = (df_clean_formatted['EUR_USD'] > 1.7).sum()

print("\n" + "="*80)
print("EUR/USD VALIDATION")
print("="*80)
print(f"Values < 0.8 (anomalous): {eur_low}")
print(f"Values > 1.7 (anomalous): {eur_high}")

if eur_low == 0 and eur_high == 0:
    print("✅ All EUR/USD values within realistic range")
else:
    print(f"⚠️ {eur_low + eur_high} anomalous EUR/USD values detected")

print("\n" + "="*80)
print("DATASET PREVIEW")
print("="*80)
print(df_clean_formatted.head(10).to_string())

print("\n" + "="*80)
print("✅ DATA PREPROCESSING COMPLETED")
print("="*80)
print(f"\nFinal dataset: {df_clean_formatted.shape[0]} months × {df_clean_formatted.shape[1]} variables")
print(f"Period: {df_clean_formatted.index.min().strftime('%Y-%m')} to {df_clean_formatted.index.max().strftime('%Y-%m')}")
print(f"Missing values: {df_clean_formatted.isnull().sum().sum()}")
print("\n⚠️ Use 'spanish_cpi_dataset_clean.xlsx' for analysis")
